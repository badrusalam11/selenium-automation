import csv
from openpyxl import load_workbook


class DataFile:
    # Read data from a CSV file
    def read_csv(file_path):
        data = []
        with open(file_path, mode='r') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                data.append(row)
        return data

    def read_excel(file_path, sheet_name):
        # Load the workbook and sheet
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]     
        # Extract headers from the first row (index 0)
        headers = [cell.value for cell in sheet[1]]
        data = []
        # Iterate over rows starting from the second row (skip the header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for each row, mapping headers to values
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            data.append(row_data)
        return data
